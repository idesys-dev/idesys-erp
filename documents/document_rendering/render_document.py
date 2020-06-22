from copy import copy
import os

from flask import flash
from flask_login import current_user
import jinja2
from docxtpl import DocxTemplate
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from template_pptx_jinja.render import PPTXRendering

from documents.models.document import Document
from external_apis.slack_bot import send_message
from external_apis.gsuite_api import credentials, gdrive

OUTPUT_DIR = "documents/document_rendering/generated_documents/"
INPUT_DIR = "documents/document_rendering/templates/"

if not os.path.exists(OUTPUT_DIR):
    os.mkdir(OUTPUT_DIR)

def render_docx(_type, data, output_name):
    """Render .docx

    The name of the variable _type begins with a _ because type in a built-in Python function.

    Example argument:
        data = { 'consultant_name' : "Jean" }
        output_name = "rm-jean"
    """
    doc = DocxTemplate(INPUT_DIR + "{_type}.docx".format(_type=_type))
    doc.render(data)
    output_path = OUTPUT_DIR + output_name + ".docx"
    doc.save(output_path)
    return output_path

def render_xlsx(_type, data, output_name):
    """Example:
        data = {
            "pay": "180",
            "jeh": "2",
        }
        output_name = 'bv-jean'
    """
    filename = INPUT_DIR + "{_type}.xlsx".format(_type=_type)
    sheetname = 'template'

    wb_template = load_workbook(filename=filename)
    ws = wb_template[sheetname]

    wb = Workbook()
    ws_out = wb.active
    ws_out.title = output_name

    for row in range(1, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            content = ws[get_column_letter(col) + str(row)].value
            if content:

                # Use french decimal format
                if isinstance(content, float):
                    content = str(content)
                    content = content.replace('.', ',')

                template = jinja2.Template(str(content))
                rendered = template.render(data)
                ws_out[get_column_letter(col) + str(row)] = rendered
                ws_out[get_column_letter(col) + str(row)].font = copy(ws[get_column_letter(col) + str(row)].font)

    output_path = OUTPUT_DIR + output_name + '.xlsx'
    wb.save(filename=output_path)
    return output_path


def render_pptx(_type, data, output_name):
    # see https://github.com/Thykof/template-pptx-jinja/blob/master/example.py
    def majuscule(input):
        return input.capitalize()

    def gender(input, value):
        # usage: {{president.sexe| gender('e')}}
        return value if input == 'f' else ''

    jinja2_env = jinja2.Environment()
    jinja2_env.filters['gender'] = gender
    jinja2_env.filters['majuscule'] = majuscule

    input_path = INPUT_DIR + "{_type}.pptx".format(_type=_type)
    output_path = OUTPUT_DIR + output_name + ".pptx"
    rendering = PPTXRendering(input_path, data, output_path, jinja2_env)
    message = rendering.process()
    print(message)  # DEBUG
    return output_path


# Define a dict to map render functions
rendering_function = {
    'docx': render_docx,
    'xlsx': render_xlsx,
    'pptx': render_pptx,
}

mime_types = {
    "docx": {
        'metadata': 'application/vnd.google-apps.document',
        'media': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    },
    "xlsx": {
        'metadata': 'application/vnd.google-apps.spreadsheet',
        'media': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    },
    "pptx": {
        'metadata': 'application/vnd.google-apps.presentation',
        'media': 'application/vnd.openxmlformats-officedocument.presentationml.presentation'
    }
}

def render_document(doc_type, file_extension, data, name):
    output_path = rendering_function[file_extension](doc_type, data, name)
    scopes = ['https://www.googleapis.com/auth/drive']
    metadata_mime_type = mime_types[file_extension]['metadata']
    media_mimetype = mime_types[file_extension]['media']
    creds = credentials.get_delegated_credentials(scopes, current_user.email)
    link = gdrive.upload(creds, output_path, metadata_mime_type, media_mimetype)
    resp = send_message.send('Nouveau document généré sur IdéSYS-ERP : ' + link, 'zapier-test')
    flash(resp)
    doc = Document(title=name, path=output_path, link=link, _type=doc_type, status="created")
    doc.save()
    return link
